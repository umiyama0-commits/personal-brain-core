"""
content_extractor.py — 共有コンテンツからテキストを抽出

対応コンテンツ:
  - Google Docs/Sheets → エクスポートURLでテキスト/CSV取得
  - 画像 → Vision API (LiteLLM経由) でOCR/内容認識
  - Excel (.xls/.xlsx) → openpyxl/xlrd でテキスト抽出
  - PDF → PyPDF2 でテキスト抽出
"""

import os
import re
import json
import base64
import logging
import tempfile
import asyncio
import ipaddress
import socket
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, urljoin

import httpx

logger = logging.getLogger(__name__)

DOWNLOAD_DIR = Path("/tmp/brain_downloads")

# ★2026-07-01 抽出/OCR 用モデル (§llm-switching-policy A/B)。Sonnet 5 は Opus 4.8 と
# テキスト抽出同等 + 数値OCRが byte 一致 (synthetic 21/21・実dashboard 全数一致) を実証、~55%減。
# .env の EXTRACT_MODEL で切替。未設定なら smart=Opus 4.8 に安全フォールバック (revert = .env の1行削除)。
EXTRACT_MODEL = os.getenv("EXTRACT_MODEL", "smart")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ★2026-05-13: 汎用 URL 取得 → テキスト抽出 (LINE 経由の URL 共有用)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# HTML 抽出のフォールバック regex (依存無しでも最低限動作させるため)
_HTML_SCRIPT_RE = re.compile(r"<(script|style|noscript)[^>]*>.*?</\1>", re.DOTALL | re.IGNORECASE)
_HTML_TAG_RE = re.compile(r"<[^>]+>")
_HTML_WS_RE = re.compile(r"[ \t]+")
_HTML_BLANK_RE = re.compile(r"\n{3,}")


def _simple_html_to_text(html: str) -> str:
    """フォールバック HTML → text 変換 (trafilatura/bs4 が無い時)"""
    import html as _html
    text = _HTML_SCRIPT_RE.sub("", html)
    text = re.sub(r"</?(br|p|div|li|h[1-6]|tr|article|section)[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = _HTML_TAG_RE.sub("", text)
    text = _html.unescape(text)
    text = _HTML_WS_RE.sub(" ", text)
    text = _HTML_BLANK_RE.sub("\n\n", text)
    return text.strip()


def _extract_main_content(html: str, url: str) -> tuple[str, str]:
    """HTML から本文 + title を取り出す。
    優先順: trafilatura (本文 + メタ抽出が最強) → readability → bs4 → 簡易 regex

    Returns: (title, body_text)
    """
    title = ""
    body = ""

    # ① trafilatura (本文抽出に特化、navigation/footer を自動除去)
    try:
        import trafilatura  # type: ignore
        extracted = trafilatura.extract(
            html,
            include_comments=False,
            include_tables=True,
            favor_recall=True,  # 多めに取る
            url=url,
        )
        if extracted:
            body = extracted
        meta = trafilatura.extract_metadata(html)
        if meta and meta.title:
            title = meta.title
    except Exception as e:
        logger.debug(f"trafilatura unavailable or failed: {e}")

    # ② readability (Mozilla Readability port) — trafilatura で取れなかったら
    if not body:
        try:
            from readability import Document  # type: ignore
            doc = Document(html)
            if not title:
                title = doc.title() or ""
            content_html = doc.summary(html_partial=True)
            # bs4 で text 化
            from bs4 import BeautifulSoup  # type: ignore
            soup = BeautifulSoup(content_html, "lxml")
            body = soup.get_text(separator="\n", strip=True)
        except Exception as e:
            logger.debug(f"readability unavailable or failed: {e}")

    # ③ bs4 直で (readability 使えない場合)
    if not body:
        try:
            from bs4 import BeautifulSoup  # type: ignore
            soup = BeautifulSoup(html, "lxml")
            if not title and soup.title:
                title = soup.title.string or ""
            for tag in soup(["script", "style", "noscript", "nav", "aside", "footer", "header"]):
                tag.decompose()
            # main / article / body の順で優先
            for selector in ["main", "article", "[role='main']", "#main", "#content", "body"]:
                node = soup.select_one(selector)
                if node:
                    body = node.get_text(separator="\n", strip=True)
                    if len(body) > 200:
                        break
        except Exception as e:
            logger.debug(f"bs4 unavailable or failed: {e}")

    # ④ フォールバック: 全 regex 除去
    if not body:
        body = _simple_html_to_text(html)

    if not title:
        m = re.search(r"<title[^>]*>(.*?)</title>", html, re.DOTALL | re.IGNORECASE)
        title = (m.group(1).strip() if m else url)

    return title.strip()[:200], body.strip()


# ─── SSRF ガード (★2026-06-08 システム評価 Security HIGH-2) ───
# extract_url は LINE Works 社員の任意メッセージ中の URL を fetch する (= 非管理者から到達可能)。
# allowlist/private-IP block が無いと http://localhost:4000 (LiteLLM 全鍵)/169.254.169.254
# (cloud metadata)/RFC1918 内部ネットをサーバに踏ませる SSRF になる。解決済み IP を検証して
# public 以外を拒否し、redirect も各 hop で再検証する (follow_redirects 任せにしない)。
_SSRF_BLOCKED_HOSTNAMES = {"metadata.google.internal", "metadata.goog"}


def _validate_public_url(url: str) -> tuple[bool, str]:
    """url の scheme/host を検証し、解決後 IP が public でなければ (False, 理由)。"""
    try:
        p = urlparse(url)
    except Exception as e:
        return False, f"unparseable url: {e}"
    if p.scheme not in ("http", "https"):
        return False, f"scheme not allowed: {p.scheme or '(none)'}"
    host = (p.hostname or "").strip()
    if not host:
        return False, "no host"
    if host.lower() in _SSRF_BLOCKED_HOSTNAMES:
        return False, f"blocked hostname: {host}"
    try:
        port = p.port or (443 if p.scheme == "https" else 80)
    except ValueError:
        return False, "invalid port"
    # host を解決し、全ての解決先 IP を検証 (DNS rebinding の主要ケースを潰す)
    try:
        infos = socket.getaddrinfo(host, port, proto=socket.IPPROTO_TCP)
    except Exception as e:
        return False, f"dns resolution failed: {e}"
    for info in infos:
        ip_str = info[4][0]
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            return False, f"invalid resolved ip: {ip_str}"
        if (ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved
                or ip.is_multicast or ip.is_unspecified):
            return False, f"non-public ip {ip_str} (host {host})"
    return True, "ok"


async def _safe_get(
    http: httpx.AsyncClient,
    url: str,
    *,
    timeout: float,
    headers: dict,
    max_redirects: int = 5,
):
    """SSRF ガード付き GET。各 hop で解決後 IP の public 性を検証してから接続。

    private/loopback/link-local/metadata を拒否。follow_redirects は使わず手動で辿り、
    redirect 先 (内部ネットへの誘導) も毎回検証する。検証は同期 getaddrinfo なので
    event loop を塞がないよう to_thread で実行。
    """
    current = url
    for _ in range(max_redirects + 1):
        ok, reason = await asyncio.to_thread(_validate_public_url, current)
        if not ok:
            raise ValueError(f"SSRF blocked: {reason}")
        resp = await http.get(current, timeout=timeout, follow_redirects=False, headers=headers)
        if resp.is_redirect and resp.headers.get("location"):
            current = urljoin(current, resp.headers["location"])
            continue
        return resp
    raise ValueError("SSRF blocked: too many redirects")


async def extract_url(
    url: str,
    http: httpx.AsyncClient,
    max_chars: int = 50000,
    timeout: float = 60.0,
) -> Optional[str]:
    """汎用 URL → text 抽出。

    ★2026-05-13 v2: trafilatura / readability / bs4 で本文抽出精度を大幅向上、
    max_chars / timeout も拡大、PDF も max_pages 80 に。

    対応:
      - HTML → trafilatura で本文中心抽出 (nav/footer 除外)
      - JSON / text/plain / markdown → そのまま
      - PDF → extract_file_text 経由 (80 ページまで)
      - Google Docs/Sheets → extract_google_doc 経由
      - その他バイナリ → エラーメッセ
    """
    # Google Docs/Sheets は専用関数で
    if "docs.google.com" in url or "drive.google.com" in url:
        try:
            txt = await extract_google_doc(url, http)
            if txt:
                return txt[:max_chars]
        except Exception as e:
            logger.warning(f"extract_google_doc failed for {url}: {e}")
            # 一般 HTML 経路にフォールバック

    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            "Accept-Language": "ja,en;q=0.9",
        }
        resp = await _safe_get(http, url, timeout=timeout, headers=headers)
        resp.raise_for_status()
    except Exception as e:
        return f"[URL fetch error: {url}\n  {type(e).__name__}: {e}]"

    content_type = (resp.headers.get("content-type", "") or "").lower()
    raw = resp.content

    if "html" in content_type or "xml" in content_type:
        try:
            html = resp.text
            title, body = _extract_main_content(html, url)
            out = f"[Web ページ] {title}\nURL: {url}\n\n{body}"
            return out[:max_chars]
        except Exception as e:
            return f"[HTML parse error: {e}]"

    if "json" in content_type:
        return f"[JSON] {url}\n{resp.text[:max_chars]}"

    if content_type.startswith("text/") or "markdown" in content_type:
        return f"[Text] {url}\n{resp.text[:max_chars]}"

    if "pdf" in content_type:
        try:
            tmp = Path(tempfile.mktemp(suffix=".pdf"))
            tmp.write_bytes(raw)
            text = await extract_file_text(tmp, max_chars=max_chars, max_pages=80)
            tmp.unlink(missing_ok=True)
            return f"[PDF] {url}\n{text or '(抽出失敗)'}"
        except Exception as e:
            return f"[PDF extract error: {e}]"

    return f"[Unsupported content-type: {content_type}] {url}"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Google Docs / Sheets
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _extract_gdoc_id(url: str) -> Optional[str]:
    """Google Docs/Sheets URLからドキュメントIDを抽出"""
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    return m.group(1) if m else None


def _get_gdoc_export_url(url: str) -> Optional[str]:
    """Google Docs/Sheets URLをエクスポートURLに変換"""
    doc_id = _extract_gdoc_id(url)
    if not doc_id:
        return None
    if "spreadsheets" in url or "sheets" in url:
        return f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv"
    elif "document" in url:
        return f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    elif "presentation" in url:
        return f"https://docs.google.com/presentation/d/{doc_id}/export?format=txt"
    return None


async def extract_google_doc(url: str, http: httpx.AsyncClient) -> Optional[str]:
    """Google Docs/Sheets の内容をテキストとして取得.

    ★2026-05-25 fix: OAuth path 優先。
    制限共有 sheet (= 「bot-account@example.co.jp に共有」) は anonymous HTTP fetch だと
    401 で失敗するため、まず services.drive_ingest.fetch_text (= OAuth Drive API)
    を試し、失敗時のみ既存 anonymous export URL fallback。
    """
    # ── OAuth 優先 (= services.drive_ingest 経由) ──
    try:
        import asyncio as _aio
        from services.drive_ingest import fetch_text as _drive_fetch
        result = await _aio.to_thread(_drive_fetch, url, 5000)
        if result and result.get("ok"):
            txt = result.get("text", "")
            logger.info(f"Google Doc OAuth 取得: {len(txt)} chars from {url[:60]}")
            return txt
        # 制限 sheet で permission_denied / not_found なら anonymous fallback も無意味
        err_code = (result or {}).get("error_code", "")
        if err_code in ("permission_denied", "not_found"):
            logger.info(f"Google Doc OAuth: {err_code} → anonymous fallback skip")
            return None
        # credentials_missing / 5xx の場合は anonymous で 公開 sheet 救済の可能性ある
    except Exception as e:
        logger.warning(f"OAuth drive_ingest 失敗 → anonymous fallback: {e}")

    # ── 既存 anonymous fallback (= 公開 sheet のみ動作) ──
    export_url = _get_gdoc_export_url(url)
    if not export_url:
        return None
    try:
        resp = await http.get(export_url, follow_redirects=True, timeout=30.0)
        if resp.status_code == 200:
            text = resp.text
            if len(text) > 5000:
                text = text[:5000] + "\n...(truncated)"
            logger.info(f"Google Doc anonymous 取得: {len(text)} chars from {url[:60]}")
            return text
        else:
            logger.warning(f"Google Doc anonymous 取得失敗: {resp.status_code} for {url[:60]}")
            return None
    except Exception as e:
        logger.warning(f"Google Doc anonymous 取得エラー: {e}")
        return None


async def extract_google_doc_via_playwright(url: str, page) -> Optional[str]:
    """Playwright経由でGoogle Docs/Sheetsの内容を取得（認証済みセッション利用）"""
    export_url = _get_gdoc_export_url(url)
    if not export_url:
        return None
    try:
        response = await page.request.get(export_url)
        if response.ok:
            text = await response.text()
            if len(text) > 5000:
                text = text[:5000] + "\n...(truncated)"
            logger.info(f"Google Doc取得(playwright): {len(text)} chars")
            return text
        return None
    except Exception as e:
        logger.warning(f"Google Doc(playwright)エラー: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 画像 → Vision API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def extract_image_text(
    image_url: str,
    http: httpx.AsyncClient,
    litellm_url: str,
    litellm_key: str,
) -> Optional[str]:
    """画像URLからVision APIでテキスト/内容を抽出"""
    try:
        resp = await http.post(
            f"{litellm_url}/v1/chat/completions",
            headers={"Authorization": f"Bearer {litellm_key}"},
            json={
                "model": EXTRACT_MODEL,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": "この画像の内容を日本語で簡潔に説明してください。テキストが含まれている場合はOCRして全文を書き起こしてください。",
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": image_url},
                            },
                        ],
                    }
                ],
                "max_tokens": 1000,
            },
            timeout=30.0,
        )
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"]
        logger.info(f"画像テキスト抽出: {len(text)} chars")
        return text
    except Exception as e:
        logger.warning(f"画像テキスト抽出エラー: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Magic byte sniffer (LINE Works が fileName を返さないとき用)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def sniff_extension(data: bytes) -> str:
    """ファイル先頭バイトから拡張子を推定。判別不能なら空文字。

    対応:
      - ZIP-based Office: .docx / .xlsx / .pptx (中身を覗く)
      - .pdf, .png, .jpg, .gif, .webp, .heic
      - 旧 Office (CFB): .doc / .xls / .ppt — まとめて .doc 扱い (xlrd で .xls だけ別 path)
      - text っぽい (UTF-8 / ASCII printable) → .txt
    """
    if not data:
        return ""

    head = data[:16]

    # ZIP container (Office Open XML はすべて ZIP)
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK\x05\x06") or head.startswith(b"PK\x07\x08"):
        try:
            import zipfile, io
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                names = set(z.namelist())
                if any(n.startswith("word/") for n in names):
                    return ".docx"
                if any(n.startswith("xl/") for n in names):
                    return ".xlsx"
                if any(n.startswith("ppt/") for n in names):
                    return ".pptx"
            return ".zip"
        except Exception:
            return ".zip"

    if head.startswith(b"%PDF"):
        return ".pdf"
    if head.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if head.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if head.startswith(b"GIF87a") or head.startswith(b"GIF89a"):
        return ".gif"
    if head.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp"
    if head[4:8] == b"ftyp" and data[8:12] in (b"heic", b"heix", b"mif1", b"msf1"):
        return ".heic"
    # ★2026-05-27 海山指示: 動画 file format 拡張 (.mov / .mp4 / .m4v / .avi / .mkv / .webm)
    # ftyp box (= ISOBMFF) で .mp4 / .mov / .m4v 系
    if head[4:8] == b"ftyp":
        brand = data[8:12]
        if brand in (b"qt  ", b"mov "):
            return ".mov"
        if brand in (b"M4V ", b"M4VH", b"M4VP", b"mp42"):
            return ".m4v"
        # mp4 系 (= isom / mp41 / mp42 / iso2 / avc1 / dash 等、汎用)
        return ".mp4"
    # AVI (= RIFF...AVI )
    if head.startswith(b"RIFF") and data[8:12] == b"AVI ":
        return ".avi"
    # WebM / Matroska (= EBML header 1A 45 DF A3)
    if head.startswith(b"\x1a\x45\xdf\xa3"):
        return ".mkv"  # WebM も同 magic、ffmpeg 側で透過処理可
    # FLV (= "FLV\x01")
    if head.startswith(b"FLV\x01"):
        return ".flv"
    # 旧 MS Office (CFB / OLE compound document)
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        # .doc / .xls / .ppt 共通シグネチャ。ここでは .doc にしておく
        # (extract_file_text 側は .xls だけ別 handler)
        return ".doc"

    # text っぽいか
    try:
        sample = data[:4096].decode("utf-8")
        # 制御文字 (TAB/LF/CR以外) が極端に少なければ text
        ctrl = sum(1 for c in sample if ord(c) < 32 and c not in "\t\n\r")
        if len(sample) > 0 and ctrl / max(len(sample), 1) < 0.05:
            return ".txt"
    except UnicodeDecodeError:
        pass

    return ""


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ファイル → テキスト抽出
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def extract_file_text(
    file_path: Path,
    max_chars: int = 5000,
    max_pages: int = 20,
    max_sheets: int = 5,
    max_rows_per_sheet: int = 100,
) -> Optional[str]:
    """ダウンロード済みファイルからテキストを抽出

    パラメータ既定値は data/brain/import/ 経由の従来用途を維持。
    うみやまAI のアップロード処理など大きなファイルを扱う場合は
    呼び出し側で max_chars/max_pages/max_sheets/max_rows_per_sheet を上げる。
    """
    suffix = file_path.suffix.lower()

    try:
        if suffix in (".xlsx", ".xls"):
            return _extract_excel(file_path, max_chars, max_sheets, max_rows_per_sheet)
        elif suffix == ".pdf":
            text = _extract_pdf(file_path, max_chars, max_pages)
            # 画像化 PDF (テキスト層なし) → Vision-OCR fallback
            if _pdf_extract_was_empty(text):
                if _should_skip_ocr_for_sales(file_path):
                    logger.info(
                        f"画像化 PDF だが売上数値系のため OCR スキップ "
                        f"(スクレイパーが権威ソース): {file_path.name}"
                    )
                    return (
                        f"{text}\n\n"
                        f"[注: 売上数値系ファイルのため Vision-OCR は意図的に未実行。"
                        f"OWNDAYS 売上は mobile.owndays.net スクレイパーが API から"
                        f"正確・完全に取得済み (owndays-daily-sales / owndays-history-*) "
                        f"が権威ソース。この PDF は冗長スナップショットのため非取込。]"
                    )
                ocr = await _ocr_pdf_with_vision(file_path, max_chars, max_pages)
                if ocr:
                    return ocr
                logger.info(f"Vision-OCR も不調、空マーカーを返す: {file_path.name}")
            return text
        elif suffix in (".csv", ".tsv", ".txt", ".md", ".log", ".json", ".yaml", ".yml"):
            text = file_path.read_text(encoding="utf-8", errors="replace")
            return text[:max_chars] if len(text) > max_chars else text
        elif suffix == ".docx":
            return _extract_docx(file_path, max_chars)
        elif suffix == ".pptx":
            return _extract_pptx(file_path, max_chars)
        else:
            logger.info(f"未対応ファイル形式: {suffix}")
            return None
    except Exception as e:
        logger.warning(f"ファイルテキスト抽出エラー ({file_path.name}): {e}")
        return None


def _extract_excel(
    file_path: Path,
    max_chars: int = 5000,
    max_sheets: int = 5,
    max_rows_per_sheet: int = 100,
) -> Optional[str]:
    """Excel ファイルからテキスト抽出"""
    # .xls (旧形式) は xlrd で処理
    if file_path.suffix.lower() == ".xls":
        return _extract_xls_legacy(file_path, max_chars, max_sheets, max_rows_per_sheet)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames[:max_sheets]:
            ws = wb[sheet]
            lines.append(f"[Sheet: {sheet}]")
            for row in ws.iter_rows(max_row=max_rows_per_sheet, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        wb.close()
        text = "\n".join(lines)
        return text[:max_chars] if len(text) > max_chars else text
    except ImportError:
        logger.warning("openpyxl not installed, trying xlrd")
        return _extract_xls_legacy(file_path, max_chars, max_sheets, max_rows_per_sheet)


def _extract_xls_legacy(
    file_path: Path,
    max_chars: int = 5000,
    max_sheets: int = 5,
    max_rows_per_sheet: int = 100,
) -> Optional[str]:
    """旧 .xls 形式を xlrd で処理"""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(file_path))
        lines = []
        for sheet in wb.sheets()[:max_sheets]:
            lines.append(f"[Sheet: {sheet.name}]")
            for row_idx in range(min(sheet.nrows, max_rows_per_sheet)):
                cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
                if any(cells):
                    lines.append("\t".join(cells))
        text = "\n".join(lines)
        return text[:max_chars] if len(text) > max_chars else text
    except ImportError:
        logger.warning("xlrd not installed, skipping .xls extraction")
        return None


def _looks_like_mojibake(text: str) -> bool:
    """日本語 PDF が化けてるか粗く判定。

    macOS Quartz が export する PDF は ToUnicode CMap が無いことがあり、
    pypdf が glyph index をそのまま char にしてしまう (Bengali / Hangul 等の
    ブロックに化ける)。化けた text は wiki に入れたくない。

    判定: 「日本語 (CJK / ひらがな / カタカナ) を 1 文字も含まず、
    BMP の Indic / Hangul / Cyrillic 等が大半を占める」場合は化けと見なす。
    ASCII のみの英文 PDF は False (= 化けではない)。
    """
    if not text or len(text) < 5:
        return False
    has_cjk = any(
        "぀" <= c <= "ヿ"  # ひらがな・カタカナ
        or "一" <= c <= "鿿"  # CJK 統合漢字
        or "＀" <= c <= "￯"  # 全角英数
        for c in text
    )
    if has_cjk:
        return False
    # 化けやすい範囲 (Bengali / Devanagari / Hangul / Cyrillic / 制御文字以外の Latin)
    suspicious = sum(
        1 for c in text
        if "ऀ" <= c <= "෿"  # Devanagari, Bengali, Tamil 等
        or "Ѐ" <= c <= "ӿ"  # Cyrillic
        or "가" <= c <= "힯"  # Hangul
        or "ʰ" <= c <= "˿"  # Spacing Modifier Letters
    )
    # ASCII (= まともな英語可能性) を除外したいので suspicious が text の半分以上なら化け
    return suspicious >= len(text.strip()) * 0.3


def _extract_pdf_with_pdfminer(file_path: Path, max_chars: int, max_pages: int) -> Optional[str]:
    """pdfminer.six で抽出。日本語 ToUnicode 無し PDF にも強い。"""
    try:
        from pdfminer.high_level import extract_text
        from pdfminer.layout import LAParams
    except ImportError:
        return None
    try:
        text = extract_text(
            str(file_path),
            page_numbers=list(range(max_pages)),
            laparams=LAParams(),
        )
        return text.strip() if text else ""
    except Exception as e:
        logger.warning(f"pdfminer extract failed: {file_path.name}: {e}")
        return None


def _extract_pdf_with_pypdf(file_path: Path, max_pages: int) -> tuple[Optional[str], int, int]:
    """pypdf で抽出。返り値: (text, n_pages, pages_with_text)。"""
    try:
        import pypdf as _pdf  # type: ignore
    except ImportError:
        try:
            import PyPDF2 as _pdf  # type: ignore
        except ImportError:
            return None, 0, 0
    try:
        reader = _pdf.PdfReader(str(file_path))
    except Exception as e:
        logger.warning(f"pypdf read failed: {file_path.name}: {e}")
        return None, 0, 0
    n_pages = len(reader.pages)
    lines: list[str] = []
    pages_with_text = 0
    for page in reader.pages[:max_pages]:
        try:
            t = page.extract_text()
        except Exception:
            continue
        if t and t.strip():
            lines.append(t)
            pages_with_text += 1
    return "\n".join(lines).strip(), n_pages, pages_with_text


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ★2026-05-19: 画像化 PDF の Vision-OCR fallback
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# pdfminer / pypdf が両方空 (= テキスト層なし、スキャン/画像化 PDF) の時だけ、
# PyMuPDF でページを画像化 → Vision LLM (smart) で書き起こす最終手段。
#
# ★信頼性ガード (CLAUDE.md「それっぽいけど怪しい内容を書かない」):
#  1. 売上数値系ファイルは OCR しない (_OCR_SKIP_SALES_RE)。OWNDAYS 売上は
#     mobile.owndays.net スクレイパーが API から完全・正確・決定論的に
#     取得済みでそれが権威ソース。OCR の桁誤読を最も精度クリティカルな
#     領域に注入するのは sales_accuracy_check が防ぐ "wrong data" そのもの。
#  2. OCR 結果は冒頭に confidence マーカーを付与 → compile / clone が
#     「OCR 由来 = 断定しない、数値・固有名詞は要原典確認」と扱える。
#  3. ページ数上限 (max_pages) でコスト/レイテンシを制御。

# OCR をスキップする「売上数値が主体」のファイル名パターン (権威ソース有)。
_OCR_SKIP_SALES_RE = re.compile(
    r"(営業数値|売上数値|売上速報|売上日報|売上実績表|月次売上|日次売上"
    r"|sales[\s_\-]?(numbers?|figures?|report|data))",
    re.IGNORECASE,
)

# _extract_pdf が画像化 PDF と判定した時に返すマーカーの接頭辞
_PDF_EMPTY_MARKER_PREFIX = "[PDF text-extract empty]"


def _pdf_extract_was_empty(text: Optional[str]) -> bool:
    """_extract_pdf の戻り値が「画像化 PDF (テキスト抽出不能)」マーカーか。"""
    return bool(text) and text.lstrip().startswith(_PDF_EMPTY_MARKER_PREFIX)


def _should_skip_ocr_for_sales(file_path: Path) -> bool:
    """売上数値系は OCR しない (スクレイパーが権威ソース、OCR は冗長+誤読リスク)。"""
    return bool(_OCR_SKIP_SALES_RE.search(file_path.name))


async def _ocr_pdf_with_vision(
    file_path: Path,
    max_chars: int,
    max_pages: int,
) -> Optional[str]:
    """画像化 PDF を PyMuPDF でページ→PNG 化 → Vision LLM (smart) で書き起こす。

    最終手段 (pdfminer / pypdf が両方空の時のみ呼ばれる)。
    戻り値は confidence マーカー付き。失敗時 None。
    """
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        logger.warning(f"PyMuPDF 未導入で Vision-OCR 不可: {e}")
        return None

    litellm_url = os.getenv("LITELLM_URL", "http://litellm:4000")
    litellm_key = os.getenv("LITELLM_MASTER_KEY", "")
    if not litellm_key:
        logger.warning("LITELLM_MASTER_KEY 未設定で Vision-OCR 不可")
        return None

    try:
        doc = fitz.open(str(file_path))
    except Exception as e:
        logger.warning(f"PyMuPDF open 失敗 {file_path.name}: {e}")
        return None

    n_pages = doc.page_count
    pages_to_do = min(n_pages, max_pages)
    out_parts: list[str] = []

    prompt = (
        "これはスキャン/画像化された業務文書の1ページです。"
        "書かれている内容を**省略せず正確に**書き起こしてください。"
        "表は markdown table で再現。数値は**読み取れたものだけ**を書き、"
        "推測で補完しない。判読できない箇所は [判読不可] と明記。"
        "見たままを忠実に。要約や解釈は不要。"
    )

    async with httpx.AsyncClient(timeout=60.0) as http:
        for i in range(pages_to_do):
            try:
                page = doc.load_page(i)
                # 2x ズーム (~288dpi 相当) で小さい文字も拾う
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                png = pix.tobytes("png")
                data_url = "data:image/png;base64," + base64.b64encode(png).decode()
                resp = await http.post(
                    f"{litellm_url}/v1/chat/completions",
                    headers={"Authorization": f"Bearer {litellm_key}"},
                    json={
                        "model": EXTRACT_MODEL,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {"type": "image_url", "image_url": {"url": data_url}},
                                ],
                            }
                        ],
                        "max_tokens": 2000,
                    },
                )
                resp.raise_for_status()
                txt = resp.json()["choices"][0]["message"]["content"].strip()
                if txt:
                    out_parts.append(f"--- p.{i + 1}/{n_pages} ---\n{txt}")
            except Exception as e:
                logger.warning(f"Vision-OCR p.{i + 1} 失敗 {file_path.name}: {e}")
                out_parts.append(f"--- p.{i + 1}/{n_pages} ---\n[OCR 失敗: {e}]")

    doc.close()
    body = "\n\n".join(out_parts).strip()
    if not body or len(body) < 20:
        logger.warning(f"Vision-OCR 実質空 {file_path.name}")
        return None

    header = (
        f"[PDF Vision-OCR 抽出 — confidence: medium、OCR 由来につき数値・"
        f"固有名詞は要原典確認。{pages_to_do}/{n_pages} ページ処理]\n\n"
    )
    full = header + body
    logger.info(
        f"Vision-OCR OK: {file_path.name} ({pages_to_do}/{n_pages}p, {len(body)} chars)"
    )
    return full[:max_chars] if len(full) > max_chars else full


def _extract_pdf(
    file_path: Path,
    max_chars: int = 5000,
    max_pages: int = 20,
) -> Optional[str]:
    """PDF からテキスト抽出。

    戦略:
    1. pdfminer.six で抽出 (日本語 ToUnicode 無し PDF に強い)
    2. pdfminer の結果が空 / 化け疑いなら pypdf にフォールバック
    3. 両方ダメなら "image-only PDF と思われる" 旨を明示メッセージで返す
       (silent skip は事故源 — 2026-04-28: 組織図 PDF が wiki に入らないのを
        ユーザが気付けなかった事案あり)
    """
    n_pages = 0
    pages_with_text = 0
    chosen_text: str = ""

    # 1) pdfminer
    pm_text = _extract_pdf_with_pdfminer(file_path, max_chars, max_pages)
    if pm_text and not _looks_like_mojibake(pm_text):
        chosen_text = pm_text

    # 2) pypdf fallback (pdfminer が空 or 化けの時)
    if not chosen_text:
        pp_text, n_pages, pages_with_text = _extract_pdf_with_pypdf(file_path, max_pages)
        if pp_text and not _looks_like_mojibake(pp_text):
            chosen_text = pp_text
        elif pp_text:
            # pypdf も化け → mojibake と判定。pdfminer の方がまだましかも (空でない場合)
            if pm_text:
                chosen_text = (
                    f"[PDF text mojibake suspected — fonts may lack ToUnicode CMap. "
                    f"raw extract: {pm_text[:200]} ...]"
                )

    if not chosen_text:
        if n_pages == 0:
            try:
                import pypdf as _pdf  # noqa: F401
                from pypdf import PdfReader
                n_pages = len(PdfReader(str(file_path)).pages)
            except Exception:
                pass
        return (
            f"[PDF text-extract empty] file='{file_path.name}', pages={n_pages or '?'}. "
            f"画像化された PDF or 暗号化 PDF の可能性。"
            f"OCR (pytesseract / Vision API) や元データ (テキスト / Notes 本体) を別途共有してください。"
        )

    if n_pages and pages_with_text and pages_with_text < n_pages:
        chosen_text = (
            f"[PDF: {pages_with_text}/{n_pages} ページからテキスト抽出 "
            f"(残りは画像 or 抽出不能)]\n\n"
            + chosen_text
        )
    return chosen_text[:max_chars] if len(chosen_text) > max_chars else chosen_text


def _extract_docx(file_path: Path, max_chars: int = 5000) -> Optional[str]:
    """Word (.docx) からテキスト抽出"""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(file_path) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = tree.findall(".//w:p", ns)
        lines = []
        for p in paragraphs:
            texts = [t.text for t in p.findall(".//w:t", ns) if t.text]
            if texts:
                lines.append("".join(texts))
        text = "\n".join(lines)
        return text[:max_chars] if len(text) > max_chars else text
    except Exception as e:
        logger.warning(f"docx extraction error: {e}")
        return None


def _extract_pptx(file_path: Path, max_chars: int = 5000) -> Optional[str]:
    """PowerPoint (.pptx) からテキスト抽出"""
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        lines = []
        with zipfile.ZipFile(file_path) as z:
            slide_files = sorted(
                [n for n in z.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")],
                key=lambda n: int(n.replace("ppt/slides/slide", "").replace(".xml", "") or "0"),
            )
            for idx, slide_name in enumerate(slide_files, 1):
                lines.append(f"[Slide {idx}]")
                with z.open(slide_name) as f:
                    tree = ET.parse(f)
                for t in tree.findall(".//a:t", ns):
                    if t.text:
                        lines.append(t.text)
        text = "\n".join(lines)
        return text[:max_chars] if len(text) > max_chars else text
    except Exception as e:
        logger.warning(f"pptx extraction error: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 画像 (バイナリ) → Vision API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 動画 frame 抽出 + Vision 分析 (★2026-05-27 海山指示)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def _probe_video_duration(video_path: Path) -> Optional[float]:
    """ffprobe で動画長 (秒) を取得. 失敗時 None.

    ★2026-05-27 海山指示「frame 抽出間隔の動的調整」用. duration が分かれば
    短い動画 = 1 秒毎、長い動画 = 5 秒毎 等、動的に間隔調整できる.
    """
    import asyncio as _asyncio
    try:
        proc = await _asyncio.create_subprocess_exec(
            "ffprobe", "-v", "error",
            "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1",
            str(video_path),
            stdout=_asyncio.subprocess.PIPE,
            stderr=_asyncio.subprocess.PIPE,
        )
        stdout, _ = await _asyncio.wait_for(proc.communicate(), timeout=10)
        if proc.returncode != 0:
            return None
        duration_str = stdout.decode("utf-8", errors="ignore").strip()
        if not duration_str:
            return None
        return float(duration_str)
    except Exception as e:
        logger.warning(f"ffprobe failed: {e}")
        return None


def _calc_video_sampling(duration_sec: Optional[float]) -> tuple[int, int]:
    """duration から (every_n_seconds, max_frames) を動的に計算.

    ★2026-05-27 海山指示「短い動画は 1 秒毎、長い動画は 5 秒毎」.

    Returns: (every_n_seconds, max_frames)
      duration <= 10s:  1 秒毎 × max 10 frames  (= 短い動画は全 frame、解像度高)
      10s < d <= 30s:   3 秒毎 × max 10 frames  (= 既存 default 維持)
      30s < d <= 60s:   5 秒毎 × max 12 frames  (= ほぼ全体 cover)
      60s < d:         duration/12 秒毎 × 12 frames (= 動画全体を 12 frames で sampling)
      None / 0:         3 秒毎 × max 10 frames (= 既存 default fallback)
    """
    if not duration_sec or duration_sec <= 0:
        return 3, 10
    if duration_sec <= 10:
        return 1, 10
    if duration_sec <= 30:
        return 3, 10
    if duration_sec <= 60:
        return 5, 12
    # 60 秒超: 動画全体を 12 frames に sampling (= 最低でも 5 秒毎)
    interval = max(5, int(duration_sec / 12))
    return interval, 12


async def extract_video_thumbnail(
    video_bytes: bytes,
    max_width: int = 640,
) -> Optional[bytes]:
    """動画の最初の 1 frame (= thumbnail) を JPEG bytes で返す.

    ★2026-05-27 海山指示「動画 thumbnail prefetch」: 受信 ack 段階で 「これ送って
    もらった動画?」 と確認できるように、download 直後の 1 frame quick preview.
    cost: 0 (= ffmpeg local only、Vision 不要)
    """
    import asyncio as _asyncio
    import tempfile
    tmpdir = Path(tempfile.mkdtemp(prefix="brain_video_thumb_"))
    try:
        video_path = tmpdir / "input.mp4"
        video_path.write_bytes(video_bytes)
        out_path = tmpdir / "thumb.jpg"
        # -ss 0.5 (= 0.5 秒目の frame、開始 0 だと黒 frame 引く risk あり)
        # -vframes 1 (= 1 frame のみ)、-q:v 3 (= 高品質)
        cmd = [
            "ffmpeg", "-y", "-ss", "0.5",
            "-i", str(video_path),
            "-vframes", "1",
            "-vf", f"scale={max_width}:-1",
            "-q:v", "3",
            str(out_path),
        ]
        proc = await _asyncio.create_subprocess_exec(
            *cmd,
            stdout=_asyncio.subprocess.DEVNULL,
            stderr=_asyncio.subprocess.PIPE,
        )
        _, stderr = await _asyncio.wait_for(proc.communicate(), timeout=15)
        if proc.returncode != 0:
            logger.warning(
                f"thumbnail ffmpeg failed rc={proc.returncode}: "
                f"{stderr.decode('utf-8', errors='ignore')[:200]}"
            )
            return None
        if not out_path.exists():
            return None
        return out_path.read_bytes()
    except Exception as e:
        logger.warning(f"extract_video_thumbnail error: {e}")
        return None
    finally:
        try:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass


async def extract_video_frames(
    video_bytes: bytes,
    every_n_seconds: int = 3,
    max_frames: int = 10,
    max_width: int = 768,
) -> list[bytes]:
    """動画 bytes から N 秒毎に静止画 frame を抽出 (= ffmpeg).

    Args:
        video_bytes: 動画ファイル binary
        every_n_seconds: 抽出間隔 (= 3 秒毎 1 frame、default)
        max_frames: 最大 frame 数 (= 10 frame で 30 秒動画 cover、cost cap)
        max_width: 各 frame の最大幅 px (= cost 削減のため、Vision low-detail には 512-768 が optimal)

    Returns: list[bytes] (= 各 frame は JPEG bytes)
    cost 見積もり: 10 frames × 250 tokens (low-detail GPT-4o) = ~2500 input tokens = ~$0.0125
    """
    import asyncio as _asyncio
    import tempfile

    frames: list[bytes] = []
    tmpdir = Path(tempfile.mkdtemp(prefix="brain_video_"))
    try:
        video_path = tmpdir / "input.mp4"
        video_path.write_bytes(video_bytes)
        # ffmpeg で N 秒毎に frame 抽出 + width scale
        # -vf "fps=1/{N},scale={W}:-1" = N 秒毎 fps + 幅 W に scale (高さ自動)
        # -frames:v {max_frames} = 上限
        out_pattern = str(tmpdir / "frame_%03d.jpg")
        cmd = [
            "ffmpeg", "-y",
            "-i", str(video_path),
            "-vf", f"fps=1/{every_n_seconds},scale={max_width}:-1",
            "-frames:v", str(max_frames),
            "-q:v", "5",  # JPEG quality (1-31、5 で良画質低 size)
            out_pattern,
        ]
        proc = await _asyncio.create_subprocess_exec(
            *cmd,
            stdout=_asyncio.subprocess.DEVNULL,
            stderr=_asyncio.subprocess.PIPE,
        )
        _, stderr = await _asyncio.wait_for(proc.communicate(), timeout=60)
        if proc.returncode != 0:
            logger.warning(f"ffmpeg failed rc={proc.returncode}: {stderr.decode('utf-8', errors='ignore')[:200]}")
            return []
        # 抽出 frame を sort して読込
        for p in sorted(tmpdir.glob("frame_*.jpg")):
            try:
                frames.append(p.read_bytes())
            except Exception as _e:
                logger.warning(f"frame read failed {p}: {_e}")
        logger.info(f"extract_video_frames: {len(frames)} frames extracted ({every_n_seconds}s 毎)")
        return frames
    except _asyncio.TimeoutError:
        logger.warning("ffmpeg timeout (60s)")
        return []
    except FileNotFoundError as e:
        logger.warning(f"ffmpeg not installed in container: {e}")
        return []
    except Exception as e:
        logger.warning(f"extract_video_frames error: {e}")
        return []
    finally:
        # cleanup tmpdir
        try:
            import shutil
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass


async def extract_video_text(
    video_bytes: bytes,
    http: httpx.AsyncClient,
    litellm_url: str,
    litellm_key: str,
    every_n_seconds: Optional[int] = None,
    max_frames: Optional[int] = None,
    model: str = "fast-gpt",
) -> Optional[str]:
    """動画 bytes → frame 抽出 → 各 frame を Vision 分析 → text concat.

    Args:
        video_bytes: 動画ファイル binary
        http / litellm_url / litellm_key: LiteLLM proxy 経由 Vision API 呼出用
        every_n_seconds: 抽出間隔 (= None なら ffprobe で duration 取得 → 動的に決定).
                          ★2026-05-27 海山指示「短い動画は 1 秒毎、長い動画は 5 秒毎」.
        max_frames: 最大 frame 数 (= None なら duration に応じて 10-12 で動的).
        model: Vision API model (= default fast-gpt = gpt-5.4-mini、cost 経済性).

    Returns: 各 frame の Vision 分析結果を timestamp + 動画 metadata 付きで concat した text、又は None
    """
    # ★2026-05-27 海山指示 動的調整: duration probe → sampling 戦略 決定
    duration_sec: Optional[float] = None
    if every_n_seconds is None or max_frames is None:
        import tempfile
        tmpdir = Path(tempfile.mkdtemp(prefix="brain_video_probe_"))
        try:
            video_path = tmpdir / "input.mp4"
            video_path.write_bytes(video_bytes)
            duration_sec = await _probe_video_duration(video_path)
        finally:
            try:
                import shutil
                shutil.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass
        auto_interval, auto_max = _calc_video_sampling(duration_sec)
        if every_n_seconds is None:
            every_n_seconds = auto_interval
        if max_frames is None:
            max_frames = auto_max
        logger.info(
            f"extract_video_text: duration={duration_sec}s "
            f"→ every_n={every_n_seconds}s, max_frames={max_frames}"
        )

    frames = await extract_video_frames(
        video_bytes,
        every_n_seconds=every_n_seconds,
        max_frames=max_frames,
    )
    if not frames:
        return None

    # ★2026-05-27 海山指示: 街並み動画 → 立地評価 用 prompt (= 通行量 / 導線 / 入口)
    prompt = (
        "この映像のシーンを 1-2 行 で日本語で記述してください。\n"
        "特に以下を観察してください (= 立地評価用):\n"
        "- 通行量 (= 多い / 普通 / 少ない)\n"
        "- 歩くスピード / 方向 (= どの向きから流れてる)\n"
        "- 物件 / 店舗の見え方 / 入口の入りやすさ\n"
        "- 周辺の業種 / 雰囲気 (= 商店街 / 路地裏 / 駅前 等)\n"
        "観察に基づく事実のみ、推測 / 評価コメントは加えない。"
    )

    # 各 frame を 並列で Vision 分析 (= asyncio.gather で速度確保、~10 frames で OK)
    import asyncio as _asyncio_mod
    async def _analyze_frame(idx: int, frame_bytes: bytes) -> str:
        try:
            text = await extract_image_bytes(
                frame_bytes, http, litellm_url, litellm_key,
                mime="image/jpeg",
                model=model,
                max_tokens=200,  # 1-2 行記述で十分
                prompt=prompt,
            )
            ts = idx * every_n_seconds  # 概算秒
            if text:
                return f"[t={ts:02d}s] {text.strip()}"
            return f"[t={ts:02d}s] (分析失敗)"
        except Exception as e:
            logger.warning(f"frame {idx} analyze failed: {e}")
            return f"[t={idx * every_n_seconds:02d}s] (例外)"

    tasks = [_analyze_frame(i, f) for i, f in enumerate(frames)]
    results = await _asyncio_mod.gather(*tasks, return_exceptions=False)
    body = "\n".join(results)
    return (
        f"【動画解析結果 (= {len(frames)} frames, {every_n_seconds}s 毎)】\n"
        f"{body}"
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 音声 → text 書き起こし (★2026-05-27 海山指示「音声会話対応」 input path)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def extract_audio_text(
    audio_bytes: bytes,
    http: httpx.AsyncClient,
    litellm_url: str,
    litellm_key: str,
    file_name: str = "audio.m4a",
    mime: str = "audio/m4a",
    language: str = "ja",
    model: str = "whisper",
) -> Optional[str]:
    """音声 bytes → LiteLLM proxy 経由 Whisper API → text 書き起こし.

    Args:
        audio_bytes: 音声 file binary (.m4a / .mp3 / .wav / .ogg / .webm 等、Whisper 対応)
        http / litellm_url / litellm_key: LiteLLM proxy 経由
        file_name: API に渡す file 名 (= 拡張子は format hint として重要)
        mime: 音声 MIME type
        language: ISO-639-1 (= "ja" 日本語)
        model: LiteLLM proxy 内の model name (= default "whisper" = openai/whisper-1)

    Returns: 書き起こし text、又は None.
    cost: Whisper API = $0.006/分 (= 1 分音声 ~$0.006)、月 100 メッセージ平均 30 秒なら $0.30/月.
    """
    try:
        url = f"{litellm_url}/v1/audio/transcriptions"
        files = {"file": (file_name, audio_bytes, mime)}
        data = {"model": model, "language": language}
        headers = {"Authorization": f"Bearer {litellm_key}"}
        resp = await http.post(url, headers=headers, files=files, data=data, timeout=120.0)
        resp.raise_for_status()
        result = resp.json()
        text = result.get("text", "") if isinstance(result, dict) else ""
        if not text:
            logger.warning(f"Whisper returned empty text: {str(result)[:200]}")
            return None
        logger.info(f"Whisper transcription: {len(text)} chars, lang={language}")
        return text.strip()
    except Exception as e:
        logger.warning(f"extract_audio_text error: {e}")
        return None


async def extract_image_bytes(
    image_bytes: bytes,
    http: httpx.AsyncClient,
    litellm_url: str,
    litellm_key: str,
    mime: str = "image/jpeg",
    model: str = EXTRACT_MODEL,
    max_tokens: int = 1500,
    prompt: Optional[str] = None,
) -> Optional[str]:
    """ダウンロード済み画像バイナリ → base64 dataURL → Vision API でテキスト/内容抽出"""
    try:
        b64 = base64.b64encode(image_bytes).decode()
        data_url = f"data:{mime};base64,{b64}"
        if prompt is None:
            prompt = (
                "この画像の内容を日本語で構造化して説明してください。"
                "テキストが含まれている場合は OCR して全文を書き起こしてください。"
                "表・グラフ・スライド・スクショなら、要点と数値を漏らさず拾ってください。"
            )
        resp = await http.post(
            f"{litellm_url}/v1/chat/completions",
            headers={"Authorization": f"Bearer {litellm_key}"},
            json={
                "model": model,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": data_url}},
                        ],
                    }
                ],
                "max_tokens": max_tokens,
            },
            timeout=60.0,
        )
        resp.raise_for_status()
        text = resp.json()["choices"][0]["message"]["content"]
        logger.info(f"画像バイナリ → Vision: {len(text)} chars")
        return text
    except Exception as e:
        logger.warning(f"画像 (バイナリ) Vision エラー: {e}")
        return None


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Playwright経由でファイルダウンロード
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

async def download_file_from_chat(page, attach_index: int) -> Optional[Path]:
    """チャット内の添付ファイルをダウンロード（プレビュータブのURLからストレージURL抽出）"""
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    try:
        import asyncio as _asyncio
        from urllib.parse import unquote, urlparse, parse_qs

        # attachをクリック → 新しいタブでプレビューが開く
        tabs_before = len(page.context.pages)
        await page.evaluate(f"""(idx) => {{
            const attaches = document.querySelectorAll('div.attach');
            if (attaches[idx]) attaches[idx].click();
        }}""", attach_index)
        await _asyncio.sleep(2)

        # 新しいタブが開いたか確認
        pages = page.context.pages
        if len(pages) <= tabs_before:
            logger.warning("プレビュータブが開きませんでした")
            return None

        preview_page = pages[-1]
        preview_url = preview_page.url

        # URLからsourceFileUrlパラメータを抽出
        parsed = urlparse(preview_url)
        params = parse_qs(parsed.query)
        file_url = params.get("sourceFileUrl", [None])[0]

        if not file_url:
            logger.warning(f"sourceFileUrl not found in: {preview_url[:100]}")
            await preview_page.close()
            return None

        file_url = unquote(file_url)
        file_name = file_url.split("/")[-1].split("?")[0]

        # Playwright のリクエストコンテキストでダウンロード（認証Cookie込み）
        response = await preview_page.request.get(file_url)
        if response.ok:
            dest = DOWNLOAD_DIR / file_name
            dest.write_bytes(await response.body())
            logger.info(f"ファイルダウンロード: {dest.name} ({dest.stat().st_size} bytes)")
            await preview_page.close()
            return dest
        else:
            logger.warning(f"ファイルダウンロード失敗: HTTP {response.status}")
            await preview_page.close()
            return None

    except Exception as e:
        logger.warning(f"ファイルダウンロード失敗: {e}")
        # 開いたタブを閉じる
        try:
            pages = page.context.pages
            if len(pages) > 1:
                await pages[-1].close()
        except:
            pass
        return None
